"""
summary_files - 汇总数据到一张表里

Author: JiangHai江海
Date： 2023/4/6
"""
from pathlib import Path
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
import xlrd
from bs4 import BeautifulSoup
import time


def sum_csv(files):
    pass


def sum_xls(files_list, start_num):
    sum_wb = openpyxl.Workbook()  # type: Workbook
    sum_sheet = sum_wb.worksheets[0]  # type: Worksheet
    sum_col, sum_row = 1, 1

    for xls_file in files_list:
        wb = xlrd.open_workbook(xls_file)
        sheet = wb.sheet_by_index(0)
        print(sheet.nrows, sheet.ncols)


def sum_xlsx(x_list, x_s_num):
    sum_wb = openpyxl.Workbook()  # type: Workbook
    sum_sheet = sum_wb.worksheets[0]  # type: Worksheet
    sum_col, sum_row = 1, 1

    for xlsx_file in x_list:
        wb = openpyxl.load_workbook(xlsx_file)  # type: Workbook
        sheet = wb.worksheets[0]  # type: Worksheet
        for row in range(x_s_num, sheet.max_row+1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, col)  # type: Cell
                sum_sheet.cell(sum_row, sum_col, cell.value)
                sum_col += 1
            sum_row += 1
            sum_col = 1  # 读完一行，列重新从1开始

    sum_wb.save("./output/summary_excel.xlsx")
    print(f"合并完毕，共合并了{sum_row}行数据--->")


def sum_html(files):
    st_time = time.time()
    sum_wb = openpyxl.Workbook()  # type: Workbook
    sum_sheet = sum_wb.worksheets[0]  # type: Worksheet
    sum_col, sum_row = 1, 1
    print("数据正在导入--->\n")

    for html_file in files:
        with open(html_file, "rb") as file:
            html = file.read()
            bs_html = BeautifulSoup(html, "html.parser")
            target_table = bs_html.find("table")
            rows = target_table.find_all("tr")
            for row in rows:
                cols = row.find_all("td")
                if len(cols) > 0:
                    for col in cols:
                        sum_sheet.cell(sum_row, sum_col, col.text)
                        sum_col += 1
                    sum_row += 1
                    if sum_row % 10000 == 0:
                        print(f"已写入{sum_row}行数据--->\n")
                    sum_col = 1
    sum_wb.save("./output/summary_excel.xlsx")
    end_time = time.time()
    print(f"合并完毕，共合并了{sum_row}行数据,用时{end_time-st_time:.2f}秒--->")


def get_files_from_folder(input_path):
    folder_src = input_path  # type: Path
    files = folder_src.glob("*.*")
    return list(files)


file_type = {
    "1": sum_csv,
    "2": sum_xls,
    "3": sum_xlsx
}


def main():
    while True:
        source_src = input("请输入源数据文件夹路径：")
        source_src = Path(source_src)
        files_list = get_files_from_folder(source_src)
        if source_src.exists():
            break
        else:
            print("路径有误，请重新输入--->\n")
    save_path = Path("./output")
    if not save_path.exists():
        Path.mkdir(save_path)
    # start_num = int(input("请输入数据开始的列："))
    sum_html(files_list)


if __name__ == '__main__':
    main()
